import cPickle as pickle
import datetime
from openpyxl import Workbook
from openpyxl import styles


# TODO - make set up a function that only happens once
# - set name is code? why sometimes list and sometimes not? could it have multiple names then?
# - account for everything being None and make variables for the data
# - make variable for num cols?


# Sets up worksheet headers
def setup(ws):
    ws["A1"].value = "Task Name"
    ws["B1"].value = "Set Part"
    ws["C1"].value = "Parent Build"
    ws["D1"].value = "Start Date"
    ws["E1"].value = "End Date"


# Gets a line from the data pickle file and adds all necessary info to a worksheet
# Necessary info - Task Name (content), Set Part (entity), Parent Build (sg_parent_build),
# Start Date (start_date), Due Date (due_date)
def insert_values(ws, row_num, item):
    ws["A" + row_num].value = item["content"]
    if isinstance(item["entity"]["code"], list):   # checks if name of the entity is a list or not
        ws["B" + row_num].value = item["entity"]["code"][0]
    else:
        ws["B" + row_num].value = item["entity"]["code"]
    if not (item["sg_parent_build"] is None):     # checks if parent build exists
        ws["C" + row_num].value = item["sg_parent_build"]["code"]
    ws["D" + row_num].value = item["start_date"]
    ws["E" + row_num].value = item["due_date"]


def main():
    with open('test_data.pkl', 'rb') as f:
        data = pickle.load(f)

    wb = Workbook()
    ws = wb.active

    setup(ws)

    now = datetime.datetime.today()
    i = 2

    for item in data:
        #print(item)
        row_num = str(i)
        insert_values(ws, row_num, item)
        # better to get due date from excel or data? probably data
        due_date = datetime.datetime.strptime(item["due_date"], '%Y-%m-%d')
        if due_date < now:
            for col in ws.iter_cols(1, 5, i):
                for cell in col:
                    cell.fill = styles.fills.PatternFill(fgColor='FF0000', fill_type='solid')
        i += 1

    wb.save("projects.xlsx")


if __name__ == '__main__':
    main()
