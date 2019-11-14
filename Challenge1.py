# Written by Dina Zaslavsky
# For LAIKA Coding Test

# This program exports data from a pickle file to Microsoft Excel. It exports the Task Name, Set Part Name, Parent
# Build Name, Start Date and End Date. It is sorted by earliest Start Date first and if a Due Date has already passed
# the entire row is shaded red in Excel

import cPickle as pickle
import datetime
from openpyxl import Workbook
from openpyxl import styles


# Sets up worksheet headers
def setup(ws):
    ws["A1"].value = "Task Name"
    ws["B1"].value = "Set Part"
    ws["C1"].value = "Parent Build"
    ws["D1"].value = "Start Date"
    ws["E1"].value = "End Date"


# Gets a line from the data pickle file and adds all necessary info to a worksheet
# Necessary info - Task Name (content), Set Part (entity), Parent Build (sg_parent_build),
# Start Date (start_date), Due Date (due_date)
# Each wanted key and value are checked for existence and correct type
def insert_values(ws, i, item):
    row_num = str(i)
    if "content" in item and item["content"] and isinstance(item["content"], str):
        ws["A" + row_num].value = item["content"]
    if "entity" in item and item["entity"] and isinstance(item["entity"], dict) and "code" in item["entity"]:
        if isinstance(item["entity"]["code"], list):  # checks if name of the entity is a list or not
            ws["B" + row_num].value = item["entity"]["code"][0]
        else:
            ws["B" + row_num].value = item["entity"]["code"]
    if "sg_parent_build" in item and item["sg_parent_build"] and isinstance(item["sg_parent_build"], dict) and \
            "code" in item["sg_parent_build"] and isinstance(item["sg_parent_build"]["code"], str):
        ws["C" + row_num].value = item["sg_parent_build"]["code"]
    if "start_date" in item and item["start_date"] and isinstance(item["start_date"], str):
        ws["D" + row_num].value = item["start_date"]
    if "due_date" in item and item["due_date"] and isinstance(item["due_date"], str):
        ws["E" + row_num].value = item["due_date"]


# Checks if the due date was earlier than today and if so, it makes that row in excel red
def past_due_date(ws, item, i):
    now = datetime.datetime.today()
    if "due_date" in item and item["due_date"] and isinstance(item["due_date"], str):
        due_date = datetime.datetime.strptime(item["due_date"], '%Y-%m-%d')
        if due_date < now:
            for col in ws.iter_cols(1, 5, i):
                for cell in col:
                    cell.fill = styles.fills.PatternFill(fgColor='FF0000', fill_type='solid')


def get_start_date(item):
    return item["start_date"]

def main():

    # Testing: First run creating_pickle_tests.py
    # Change one testing boolean to True per run
    # Excel file should have empty slots where the data in the pickle file was incorrect
    testing_empty = False      # test for empty pickle file
    testing_keys = False       # test for the keys to be as expected
    testing_types = False      # test for the types to be as expected (always string except for set part may be a list)
    testing_values = False     # test for if values are missing
    testing_unordered = False  # test for if the data is not in chronological order based on start date
    if testing_empty:
        file_name = "test_data_empty.pkl"
    elif testing_keys:
        file_name = "test_data_incorrect_key.pkl"
    elif testing_types:
        file_name = "test_data_unexpected_type.pkl"
    elif testing_values:
        file_name = "test_data_values_missing.pkl"
    elif testing_unordered:
        file_name = "test_data_unordered.pkl"
    else:
        file_name = "test_data.pkl"

    # open data pickle file
    try:
        with open(file_name, 'rb') as f:
            data = pickle.load(f)
    except:
        print("ERROR: Test Data Pickle File Not Found or Empty")
        return

    data.sort(key=get_start_date)   # sort the data based on start date

    # create excel worksheet
    wb = Workbook()
    ws = wb.active

    setup(ws)

    # inputs all wanted values into worksheet and then fills in rows past due date with red
    i = 2
    for item in data:
        insert_values(ws, i, item)
        past_due_date(ws, item, i)
        i += 1

    wb.save("projects.xlsx")


if __name__ == '__main__':
    main()
